import tkinter as tk
import customtkinter as ctk
from tkinter import Toplevel, messagebox, ttk
from tkcalendar import DateEntry
from tkcalendar import Calendar
from datetime import datetime
from db_connection import connect_db
import openpyxl
from fpdf import FPDF

def show_daily_appointments(frame, staff_id):
    from home_page import create_home_page

    # Define placeholders and error labels dictionary
    error_labels = {}
    appointments = None
    staff_options = []
    staff_placeholder = "Select a staff.."
    appointment_types = ["All", "Check-Up", "Follow-Up", "Consultation"]
    appt_type_placeholder = "Select a appointment type.."

    def open_calendar(entry_widget):
        try:
            calendar_window = Toplevel()
            calendar_window.title("Select Date")
            cal = Calendar(calendar_window, selectmode="day", date_pattern="y-mm-dd")
            cal.pack(padx=10, pady=10)

            def select_date():
                entry_widget.delete(0, tk.END)
                entry_widget.insert(0, cal.get_date())
                calendar_window.destroy()

            select_button = tk.Button(calendar_window, text="Select", command=select_date)
            select_button.pack(pady=5)
        except Exception as e:
            print(f"Error opening calendar: {e}")
            messagebox.showerror("Error", "Failed to open the calendar widget.")

    # Initialize CustomTkinter appearance
    ctk.set_appearance_mode("light")  # Change to "dark" for dark mode
    ctk.set_default_color_theme("blue")  # Customize the color theme if desired

    # Clear existing widgets
    for widget in frame.winfo_children():
        widget.destroy()

    # Set up the main frame and configure the layout weights
    frame.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure(0, weight=1)

    # Main view frame
    view_frame = ctk.CTkFrame(frame, fg_color='#f0f4f8', corner_radius=10, width=600)
    view_frame.pack(fill="both", expand=True)
    view_frame.grid_columnconfigure(0, weight=1)
    view_frame.grid_rowconfigure(0, weight=1)
    view_frame.grid_rowconfigure(1, weight=1)
    view_frame.grid_rowconfigure(2, weight=1)
    view_frame.grid_rowconfigure(3, weight=1)
    view_frame.grid_rowconfigure(4, weight=1)

    # Header Frame
    header_frame = ctk.CTkFrame(view_frame, fg_color="#3b8ed0", corner_radius=10)
    header_frame.grid(row=0, column=0, columnspan=2, sticky="ew")
    header_frame.grid_columnconfigure(1, weight=1) 
    header_frame.grid_rowconfigure(0, weight=1)

    # Header Logo
    # logo = tk.PhotoImage(file="logos/daily_appointment.png")
    # logo_label = tk.Label(header_frame, image=logo, bg="#1565C0")
    # logo_label.image = logo  # Keep a reference to avoid garbage collection
    # logo_label.grid(row=0, column=0, padx=10)

    # Header Title
    title_label = ctk.CTkLabel(header_frame, text="Daily Appointment View", font=ctk.CTkFont("Helvetica", 24, "bold"), text_color="white")
    title_label.grid(row=0, column=1, pady=10, sticky="nsew")
    
    # Adjust fields_frame to allow expansion
    filter_frame = ctk.CTkFrame(view_frame, fg_color="#f0f4f8", width=600)
    filter_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=(20,0))
    filter_frame.grid_columnconfigure(0, weight=1)
    filter_frame.grid_columnconfigure(1, weight=1)
    filter_frame.grid_rowconfigure(0, weight=1)
    filter_frame.grid_rowconfigure(1, weight=1)
    filter_frame.grid_rowconfigure(2, weight=1)
    filter_frame.grid_rowconfigure(3, weight=1)
    filter_frame.grid_rowconfigure(4, weight=1)

    # Load appointments based on filters
    def load_appointments():
        global appointments
        # appointment_list.delete(0, tk.END)
        selected_date = appointment_date.get()
        search_value = search_entry.get().strip()

        appt_type_filter = appt_type_combo.get()
        if appt_type_filter == appt_type_placeholder :
            appt_type_filter = None
        else:
         appt_type_filter = appt_type_combo.get().split(" - ")[0]

        staff_filter = staff_combo.get()
        if staff_filter == staff_placeholder:
            staff_filter = None
        else:
            staff_filter = staff_combo.get().split(" - ")[0]

        if selected_date or search_value or appt_type_filter or staff_filter:    
            conn = connect_db()
            if conn:
                cursor = conn.cursor()
                # Base query
                query = """
                    SELECT a.appointment_id, CONCAT(p.first_name, ' ', p.last_name) AS patient_name,
                        a.appointment_date, a.appointment_time, a.appt_type, a.status, CONCAT(s.first_name, ' ', s.last_name) AS staff_name
                    FROM appointments a
                    JOIN patients p ON a.patient_id = p.patient_id
                    LEFT JOIN staff s ON a.staff_id = s.staff_id
                    WHERE 1=1
                """
                parameters = []

                # Add filters conditionally
                if selected_date:
                    query += "AND a.appointment_date = %s"
                    parameters.append(selected_date)
                
                if appt_type_filter == "All":
                    appt_type_filter = None

                if appt_type_filter:
                    query += " AND a.appt_type = %s"
                    parameters.append(appt_type_filter)

                if staff_filter == "All":
                    staff_filter = None
                
                if staff_filter:
                    query += " AND s.staff_id = %s"
                    parameters.append(staff_filter)
                
                if search_value:
                    query += " AND CONCAT(p.first_name, ' ', p.last_name) LIKE %s"
                    parameters.append(f"%{search_value.lower()}%") 

                cursor.execute(query, parameters)
                appointments = cursor.fetchall()
                after_appt_search_view_btn_status = "normal" if appointments else "disabled"
                after_appt_search_view_btn_color = "#3b8ed0" if appointments else "grey"
                view_details_btn.configure(state=after_appt_search_view_btn_status, fg_color=after_appt_search_view_btn_color)
                conn.close()

                # Clear previous data
                for item in appointment_tree.get_children():
                    appointment_tree.delete(item)

                # Insert new data
                if appointments:
                    for appointment in appointments:
                        appointment_tree.insert("", "end", values=appointment)
                else:
                    # Insert a placeholder row indicating no data
                    appointment_tree.insert("", "end", values=("No appointments found", "", "", "", "", ""))
        else:
            messagebox.showwarning("Select Filter", "Please select any of the filter to load the appointment.")

    # Search by Name
    se_label = ctk.CTkLabel(filter_frame, text="Search by Name:", font=ctk.CTkFont(size=12), text_color="black")
    se_label.grid(row=0, column=0, padx=25, pady=(10,0), sticky="w")
    search_entry = ctk.CTkEntry(filter_frame, font=ctk.CTkFont(size=12), placeholder_text="Enter Patient Name..", width=300, height=25)
    search_entry.grid(row=1, column=0, padx=25, pady=(0,20), sticky="w")
   
    # Select Appointment Date
    appointment_date_label = ctk.CTkLabel(filter_frame, text="Select Date:", font=ctk.CTkFont(size=12), text_color="black")
    appointment_date_label.grid(row=2, column=0, padx=25, pady=(10,0), sticky='w')
    appt_date_frame = ctk.CTkFrame(filter_frame, fg_color="#f0f4f8", width=600)
    appt_date_frame.grid(row=3, column=0, pady=(0,20), padx=25, sticky="w")
    appt_date_frame.grid_rowconfigure(0, weight=1)
    appt_date_frame.grid_columnconfigure(0, weight=1)
    appt_date_frame.grid_columnconfigure(1, weight=1)
    appointment_date = ctk.CTkEntry(appt_date_frame, placeholder_text="YYYY-MM-DD")
    appointment_date.grid(row=0, column=0, pady=(0,20), sticky="w")
    open_calendar_button = ctk.CTkButton(appt_date_frame, text="Pick a Date", command=lambda: open_calendar(appointment_date))
    open_calendar_button.grid(row=0, column=1, padx=20, pady=(0,20), sticky="w")
    
    def reset_appt_type_dropdown():
        # Restore the placeholder when the field is empty
        if not appt_type_combo.get():
            appt_type_combo.set(appt_type_placeholder)
        elif appt_type_combo.get() == appt_type_placeholder:
            appt_type_combo.set("")

        appt_type_combo.configure(values=appointment_types)  # Ensure placeholder is the first value

    def check_appt_type_input(*args):
        # global time_slots
        value = appointment_type_search_var.get()  # Get current user input
        if value == appt_type_placeholder:  # Ignore placeholder
            appt_type_combo.configure(values=appointment_types)
            return

    # Filter by Type
    at_type_label = ctk.CTkLabel(filter_frame, text="Filter by Type:", font=ctk.CTkFont(size=12), text_color="black")
    at_type_label.grid(row=0, column=1, padx=25, pady=(10,0), sticky='w') 
    appointment_type_search_var = tk.StringVar()
    appointment_type_search_var.trace_add("write", check_appt_type_input)
    appt_type_combo = ctk.CTkComboBox(filter_frame, variable=appointment_type_search_var, values=[appt_type_placeholder] + appointment_types, width=300, height=25)
    appt_type_combo.grid(row=1, column=1, padx=25, pady=(0,20), sticky="w")
    appt_type_combo.set(appt_type_placeholder)
    appt_type_combo.bind("<FocusIn>", lambda e: reset_appt_type_dropdown())
    appt_type_combo.bind("<FocusOut>", lambda e: reset_appt_type_dropdown())

    # Load staff into the dropdownx
    def load_staff():
        global staff_options
        conn = connect_db()
        if conn:
            cursor = conn.cursor()
            cursor.execute("SELECT staff_id, CONCAT(first_name, ' ', last_name, ' - ', role) FROM staff")
            staff_members = cursor.fetchall()
            staff_options = ["All"] + [f"{staff[0]} - {staff[1]}" for staff in staff_members]
            staff_combo.configure(values=staff_options)
            conn.close()

    def reset_staff_dropdown():
        global staff_options
        
        # Restore the placeholder when the field is empty
        if not staff_combo.get():
            staff_combo.set(staff_placeholder)
        elif staff_combo.get() == staff_placeholder:
            staff_combo.set("")

        staff_combo.configure(values=staff_options)  # Ensure placeholder is the first value
            
    def check_staff_input(*args):
        global staff_options

        value = staff_search_var.get()  # Get current user input
        if value == staff_placeholder:  # Ignore placeholder
            staff_combo.configure(values=staff_options)
            return
        filtered_options = [option for option in staff_options if value.lower() in option.lower()]
        staff_combo.configure(values=filtered_options)  # Add placeholder back

    # Filter by Staff
    staff_name_label = ctk.CTkLabel(filter_frame, text="Filter by Staff:", font=ctk.CTkFont(size=12), text_color="black")
    staff_name_label.grid(row=2, column=1, padx=25, pady=(10,0), sticky='w')
    staff_search_var = tk.StringVar()
    staff_search_var.trace_add("write", check_staff_input)
    staff_combo = ctk.CTkComboBox(filter_frame, variable=staff_search_var, values=staff_options, width=300, height=25)
    staff_combo.grid(row=3, column=1, padx=25, pady=(0,20), sticky="w")
    load_staff()
    staff_combo.set(staff_placeholder)
    staff_combo.bind("<FocusIn>", lambda e: reset_staff_dropdown())
    staff_combo.bind("<FocusOut>", lambda e: reset_staff_dropdown())


    load_btn = ctk.CTkButton(filter_frame, text="Load Appointments", command=load_appointments)
    load_btn.grid(row=4, column=0, columnspan=2, padx=10, pady=20)

    # Horizontal Separator using CTkFrame
    separator = ctk.CTkFrame(view_frame, height=2, fg_color="#E0E0E0", corner_radius=0)
    separator.grid(row=2, column=0, columnspan=4, sticky="ew", padx=20, pady=(20, 20))

    # Frame for appointments
    listbox_frame = ctk.CTkFrame(view_frame, fg_color="#f0f4f8", corner_radius=10)
    listbox_frame.grid(row=3, column=0, columnspan=2, padx=50, pady=(0,10), sticky="nsew")

    # Configure column and row weights for the frame
    listbox_frame.grid_columnconfigure(0, weight=1)
    listbox_frame.grid_rowconfigure(0, weight=1)

    # Treeview for displaying appointments
    columns = ("Appointment ID", "Patient Name", "Appointment Date", "Time", "Type", "Status")
    appointment_tree = ttk.Treeview(
        listbox_frame,
        columns=columns,
        show="headings",  # Hide the default empty column
        height=10,  # Adjust as needed
        padding=2
    )

    # Set up column headings
    appointment_tree.heading("Appointment ID", text="Appointment ID")
    appointment_tree.heading("Patient Name", text="Patient Name")
    appointment_tree.heading("Appointment Date", text="Appointment Date")
    appointment_tree.heading("Time", text="Time")
    appointment_tree.heading("Type", text="Type")
    appointment_tree.heading("Status", text="Status")

    # Configure column widths (optional)
    appointment_tree.column("Appointment ID", width=120, anchor="center")
    appointment_tree.column("Patient Name", width=180, anchor="center")
    appointment_tree.column("Appointment Date", width=100, anchor="center")
    appointment_tree.column("Time", width=100, anchor="center")
    appointment_tree.column("Type", width=100, anchor="center")
    appointment_tree.column("Status", width=100, anchor="center")

    # Scrollbar for the Treeview
    tree_scrollbar = ttk.Scrollbar(listbox_frame, orient="vertical", command=appointment_tree.yview)
    appointment_tree.configure(yscrollcommand=tree_scrollbar.set)

    # Grid layout for Treeview and Scrollbar
    appointment_tree.grid(row=0, column=0, sticky="nsew")
    tree_scrollbar.grid(row=0, column=1, sticky="ns")

    # Clear all fields
    def clear_fields():
        global appointments
        global staff_options
        # Reset CTkComboBox widgets to their placeholder
        staff_combo.set(staff_placeholder)
        staff_combo.configure(values=staff_options)
        appt_type_combo.set(appt_type_placeholder)
        appt_type_combo.configure(values=appointment_types)
        search_entry.delete(0, ctk.END)
        appointment_date.delete(0, ctk.END)
        appointments = None
        after_clear_field_view_btn_status = "normal" if appointments else "disabled"
        after_clear_field_view_btn_color = "#3b8ed0" if appointments else "grey"
        view_details_btn.configure(state=after_clear_field_view_btn_status, fg_color=after_clear_field_view_btn_color)

        # Clear all items in the Treeview
        for item in appointment_tree.get_children():
            appointment_tree.delete(item)
        
    # View appointment details
    def view_appointment_details():
        selected = appointment_tree.selection()
        if not selected:
            messagebox.showwarning("Select Appointment", "Please select an appointment to view details.")
            return

        selected_text = appointment_tree.item(selected[0], "values")

        headers = ["Appointment ID", "Patient", "Appointment Date", "Time", "Type", "Status", "Staff"]

        # Format or display the values as needed
        if selected_text:
            details_text = "\n".join(f"{col}: {val}" for col, val in zip(headers, selected_text))
            messagebox.showinfo("Appointment Details", details_text)
        else:
            messagebox.showerror("Error", "Could not retrieve appointment details.")

    def go_back():
        view_frame.destroy()
        create_home_page(frame, staff_id)

    def get_staff_details(staff_id):
        # Establish a database connection
        conn = connect_db()
        if conn:
            try:
                cursor = conn.cursor()

                # SQL query to get staff details based on staff_id
                cursor.execute("""
                    SELECT first_name, last_name, role
                    FROM staff
                    WHERE staff_id = %s
                """, (staff_id,))
                
                # Fetch the staff details
                staff_details = cursor.fetchone()

                if staff_details:
                    first_name, last_name, role = staff_details
                    # Return staff details
                    return first_name, last_name, role
                else:
                    return None

            except Exception as e:
                print(f"Error: {e}")
                return None
            finally:
                conn.close()
        else:
            print("Database connection error.")
            return None


    def export_to_excel():
        staff_details = get_staff_details(staff_id)

        if staff_details:
            first_name, last_name, role = staff_details
            staff_name = f"{first_name} {last_name}"
        else:
            print("Staff not found.")
            return

        # Generate file name based on current timestamp and staff name
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        file_name = f"Daily_Appointments_{first_name}_{last_name}_{current_time.replace(':', '-')}.xlsx"

        # Retrieve all rows from the Treeview
        appointments_data = []
        for item_id in appointment_tree.get_children():
            appointments_data.append(appointment_tree.item(item_id, "values"))

        if not appointments_data:
            messagebox.showerror("No Data", "No appointments to export.")
            return

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Appointments"

        # Add current time and staff name in the first row
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)  # Merge across 7 columns
        sheet.cell(row=1, column=1).value = f"Exported on: {current_time} | Staff: {staff_name}"
        sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="center")

        # Add headers
        headers = ["Appointment ID", "Patient Name", "Appointment Date", "Appointment Time", "Type", "Status", "Staff Name"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=2, column=col_num, value=header)
            sheet.cell(row=2, column=col_num).font = openpyxl.styles.Font(bold=True)

        # Add appointment data
        for row_num, appointment in enumerate(appointments_data, start=3):  # Start from row 3
            for col_num, value in enumerate(appointment, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Save the workbook
        workbook.save(file_name)
        messagebox.showinfo("Export Successful", f"Appointments exported to {file_name}")

    def export_to_pdf():
        staff_details = get_staff_details(staff_id)

        if staff_details:
            first_name, last_name, role = staff_details
            staff_name = f"{first_name} {last_name}"
        else:
            print("Staff not found.")
            return

        # Generate file name
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        file_name = f"Daily_Appointments_{first_name}_{last_name}_{current_time.replace(':', '-')}.pdf"

        # Retrieve data from Treeview
        appointments_data = []
        for item_id in appointment_tree.get_children():
            appointments_data.append(appointment_tree.item(item_id, "values"))

        if not appointments_data:
            messagebox.showerror("No Data", "No appointments to export.")
            return

        # Initialize PDF
        pdf = FPDF(format=(300, 210))
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        
        # Add header with timestamp and staff name
        pdf.set_font("Arial", "B", 12)
        pdf.cell(200, 10, txt=f"Exported on: {current_time} | Staff: {staff_name}", ln=True, align="L")
        pdf.ln(10)  # Add a blank line after the header

        # Add title
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, txt="Appointments", ln=True, align="C")
        pdf.ln(10)

        # Add headers for the table
        headers = ["Appointment ID", "Patient Name", "Appointment Date", "Appointment Time", "Type", "Status", "Staff Name"]
        column_widths = [40, 40, 40, 40, 40, 40, 40]
        row_height = 8

        pdf.set_font("Arial", "B", 10)
        for header, width in zip(headers, column_widths):
            pdf.cell(width, row_height, header, border=1, align="C")
        pdf.ln()

        # Add rows
        pdf.set_font("Arial", size=10)
        for appointment in appointments_data:
            max_line_count = 1
            for value, width in zip(appointment, column_widths):
                lines = pdf.get_string_width(str(value)) // width
                max_line_count = max(max_line_count, lines + 1)

            # Print each row, adjusting height for multi-line content
            for value, width in zip(appointment, column_widths):
                pdf.cell(width, row_height * max_line_count, str(value), border=1, align="C")
            pdf.ln(row_height * max_line_count)  # Move to the next row after the current appointment

        # Save the PDF
        pdf.output(file_name)
        messagebox.showinfo("Export Successful", f"Appointments exported to {file_name}")
        
    # Button Actions and Frame
    button_frame = ctk.CTkFrame(view_frame, fg_color="#f0f4f8")
    button_frame.grid(row=4, column=0, columnspan=2, pady=20, padx=10)
    button_frame.grid_columnconfigure(0, weight=1)
    button_frame.grid_columnconfigure(1, weight=1)
    button_frame.grid_columnconfigure(2, weight=1)
    button_frame.grid_columnconfigure(3, weight=1)
    button_frame.grid_columnconfigure(4, weight=1)
    button_frame.grid_rowconfigure(0, weight=1)

    view_btn_status = "normal" if appointments else "disabled"
    view_btn_color = "#3b8ed0" if appointments else "grey"
    view_details_btn = ctk.CTkButton(button_frame, text="View Details", command=view_appointment_details, fg_color=view_btn_color, state=view_btn_status)
    view_details_btn.grid(row=0, column=0, padx=10)

    export_to_excel_btn = ctk.CTkButton(button_frame, text="Export to Excel", command=export_to_excel)
    export_to_excel_btn.grid(row=0, column=1, padx=10)

    export_to_pdf_btn = ctk.CTkButton(button_frame, text="Export to PDF", command=export_to_pdf)
    export_to_pdf_btn.grid(row=0, column=2, padx=10)

    clear_btn = ctk.CTkButton(button_frame, text="Clear", command=clear_fields)
    clear_btn.grid(row=0, column=3, padx=10)

    back_btn = ctk.CTkButton(button_frame, text="Back", command=go_back)
    back_btn.grid(row=0, column=4, padx=10)